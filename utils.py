import os
import logging
import pandas as pd
from openpyxl import load_workbook


# def save_result(df, sheet_name):
#     """
#     Save the test results DataFrame to an Excel file.

#     Args:
#         df: pandas DataFrame with test results.
#         sheet_name: Name of the Excel sheet for this test case.
#     """
#     directory = "reports"
#     file_path = os.path.join(directory, "test_report.xlsx")

#     # Ensure the reports directory exists
#     os.makedirs(directory, exist_ok=True)

#     if os.path.exists(file_path):
#         # Open the existing workbook and check for the sheet
#         with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#             try:
#                 workbook = load_workbook(file_path)
#                 if sheet_name in workbook.sheetnames:
#                     logging.info(f"Appending to existing sheet: {sheet_name}")
#                 else:
#                     logging.info(f"Creating a new sheet: {sheet_name}")
#                     df.to_excel(writer, sheet_name=sheet_name, index=False)
#             except Exception as e:
#                 logging.error(f"Error while saving to the Excel file: {str(e)}")
#     else:
#         # Create a new workbook and add the sheet
#         with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
#             logging.info(f"Creating a new Excel file with sheet: {sheet_name}")
#             df.to_excel(writer, sheet_name=sheet_name, index=False)

#     logging.info(f"Test results saved to {file_path}")

def save_result(df, sheet_name):
    """
    Save the test results DataFrame to an Excel file.

    Args:
        df: pandas DataFrame with test results.
        sheet_name: Name of the Excel sheet for this test case.
    """
    directory = "reports"
    file_path = os.path.join(directory, "test_report.xlsx")

    # Ensure the reports directory exists
    os.makedirs(directory, exist_ok=True)

    try:
        if os.path.exists(file_path):
            # Open the existing workbook
            workbook = load_workbook(file_path)
            
            # If the sheet exists and is the only one, add a temporary sheet first
            if sheet_name in workbook.sheetnames:
                logging.info(f"Overriding existing sheet: {sheet_name}")
                if len(workbook.sheetnames) == 1:
                    workbook.create_sheet("TempSheet")
                
                # Remove the sheet to be replaced
                del workbook[sheet_name]
                workbook.save(file_path)  # Save the workbook after removing the sheet
            
            # Remove temporary sheet if it exists
            if "TempSheet" in workbook.sheetnames:
                del workbook["TempSheet"]
                workbook.save(file_path)

        # Write the new data to the specified sheet (whether file existed or not)
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a" if os.path.exists(file_path) else "w") as writer:
            logging.info(f"Saving data to sheet: {sheet_name}")
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    except Exception as e:
        logging.error(f"Error while saving to the Excel file: {str(e)}")

    logging.info(f"Test results saved to {file_path}")
